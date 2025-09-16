import shutil
import os
import pandas as pd
from openpyxl import load_workbook


origen_25_26 = os.path.expanduser('~/Downloads/Soccer-Stats-La-Liga-2025-2026_R01.xlsx')
destí_25_26 = os.path.expanduser('~/Desktop/Treball-de-Recerca/BDD_EntrenamentModel_Estadístiques-La_Liga-2025-2026.xlsx')


shutil.move(origen_25_26, destí_25_26)


wb = load_workbook(destí_25_26)

fulles_eliminar = ["La Liga", "AllFixturesExport", "TeamRosterExport", 
                   "PlayerStatsExport", "LineUpExport", "PlaysExport"]
for sheet in fulles_eliminar:
    if sheet in wb.sheetnames:
        wb.remove(wb[sheet])

canvis_noms = {"LeagueTableExport": "ClassificacióGeneral",
               "TeamStatsExport": "StatsPartit"}
for old_name, new_name in canvis_noms.items():
    if old_name in wb.sheetnames:
        wb[old_name].title = new_name

wb.save(destí_25_26)


df_25_26_classificacio = pd.read_excel(destí_25_26, sheet_name='ClassificacióGeneral')
df_25_26_partits = pd.read_excel(destí_25_26, sheet_name='StatsPartit')


df_25_26_classificacio['Team'] = df_25_26_classificacio['Team'].replace({
    'AtlÃ©tico Madrid': 'Atlético Madrid',
    'AlavÃ©s': 'Alavés'
})


df_25_26_classificacio = df_25_26_classificacio.iloc[:, 1:-1]


with pd.ExcelWriter(destí_25_26, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_25_26_classificacio.to_excel(writer, sheet_name='ClassificacióGeneral', index=False)
    df_25_26_partits.to_excel(writer, sheet_name='StatsPartit', index=False)

